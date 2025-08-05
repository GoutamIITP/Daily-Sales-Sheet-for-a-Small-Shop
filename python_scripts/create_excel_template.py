"""
Excel Template Creator for Daily Sales Sheet
Creates and formats Excel templates with formulas and validation rules.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, LineChart
from datetime import datetime, timedelta
import os


class SalesSheetCreator:
    """Creates Excel templates for daily sales tracking."""
    
    def __init__(self):
        self.workbook = Workbook()
        self.template_path = "excel_templates/daily_sales_sheet.xlsx"
        
    def create_sales_entry_sheet(self) -> None:
        """Create the main sales entry worksheet."""
        # Remove default sheet and create new one
        self.workbook.remove(self.workbook.active)
        ws = self.workbook.create_sheet("Sales Entry", 0)
        
        # Define headers
        headers = [
            "Date", "Product Name", "Category", "Quantity Sold", 
            "Unit Price", "Total Amount", "Payment Method", "Customer Type"
        ]
        
        # Add headers to row 1
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Set column widths
        column_widths = [12, 20, 15, 12, 12, 15, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        
        # Add formulas for Total Amount (column F)
        for row in range(2, 102):  # 100 rows for data entry
            ws.cell(row=row, column=6, value=f"=D{row}*E{row}")
        
        # Add data validation
        self._add_data_validation(ws)
        
        # Add borders
        self._add_borders(ws, len(headers), 101)
        
    def create_daily_summary_sheet(self) -> None:
        """Create daily summary worksheet with pivot-like functionality."""
        ws = self.workbook.create_sheet("Daily Summary")
        
        # Summary headers
        ws.cell(row=1, column=1, value="Daily Sales Summary").font = Font(size=16, bold=True)
        ws.merge_cells("A1:G1")
        
        # Summary table headers
        summary_headers = ["Date", "Total Sales", "Total Transactions", "Average Sale", "Best Selling Product", "Category Leader", "Payment Method"]
        for col, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Add formulas for daily summary
        for row in range(4, 34):  # 30 days of summary
            date_cell = ws.cell(row=row, column=1, value=f"2025-08-{row-3:02d}")
            
            # Total Sales formula (sum from Sales Entry sheet)
            ws.cell(row=row, column=2, value=f'=SUMIF(\'Sales Entry\'.A:A,A{row},\'Sales Entry\'.F:F)')
            
            # Total Transactions
            ws.cell(row=row, column=3, value=f'=COUNTIF(\'Sales Entry\'.A:A,A{row})')
            
            # Average Sale
            ws.cell(row=row, column=4, value=f'=IF(C{row}>0,B{row}/C{row},0)')
        
        # Set column widths
        for col in range(1, 8):
            ws.column_dimensions[ws.cell(row=3, column=col).column_letter].width = 15
    
    def create_product_analysis_sheet(self) -> None:
        """Create product analysis worksheet."""
        ws = self.workbook.create_sheet("Product Analysis")
        
        # Headers
        ws.cell(row=1, column=1, value="Product Performance Analysis").font = Font(size=16, bold=True)
        ws.merge_cells("A1:F1")
        
        analysis_headers = ["Product Name", "Total Sold", "Total Revenue", "Avg Price", "Last Sale Date", "Category"]
        for col, header in enumerate(analysis_headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        
        # Sample product analysis formulas (will be populated by Python script)
        products = ["Coffee", "Sandwich", "Pastry", "Juice", "Salad", "Tea", "Muffin", "Water"]
        for i, product in enumerate(products, 4):
            ws.cell(row=i, column=1, value=product)
            # Total Sold
            ws.cell(row=i, column=2, value=f'=SUMIF(\'Sales Entry\'.B:B,A{i},\'Sales Entry\'.D:D)')
            # Total Revenue
            ws.cell(row=i, column=3, value=f'=SUMIF(\'Sales Entry\'.B:B,A{i},\'Sales Entry\'.F:F)')
            # Average Price
            ws.cell(row=i, column=4, value=f'=IF(B{i}>0,C{i}/B{i},0)')
    
    def create_charts_sheet(self) -> None:
        """Create a sheet for charts and visualizations."""
        ws = self.workbook.create_sheet("Charts & Reports")
        
        ws.cell(row=1, column=1, value="Sales Charts and Visualizations").font = Font(size=16, bold=True)
        ws.merge_cells("A1:H1")
        
        # Add chart placeholders and instructions
        ws.cell(row=3, column=1, value="Chart Instructions:")
        ws.cell(row=4, column=1, value="1. Use Python scripts to generate advanced charts")
        ws.cell(row=5, column=1, value="2. Insert pivot charts for dynamic analysis")
        ws.cell(row=6, column=1, value="3. View trend analysis in visualizations folder")
    
    def _add_data_validation(self, ws) -> None:
        """Add data validation rules to the worksheet."""
        # Date validation
        date_validation = DataValidation(type="date", operator="between", 
                                       formula1="2025-01-01", formula2="2025-12-31")
        date_validation.add(f"A2:A101")
        ws.add_data_validation(date_validation)
        
        # Category validation
        categories = '"Food,Beverage,Snack,Dessert,Other"'
        category_validation = DataValidation(type="list", formula1=categories)
        category_validation.add("C2:C101")
        ws.add_data_validation(category_validation)
        
        # Payment method validation
        payment_methods = '"Cash,Credit Card,Debit Card,Mobile Payment,Other"'
        payment_validation = DataValidation(type="list", formula1=payment_methods)
        payment_validation.add("G2:G101")
        ws.add_data_validation(payment_validation)
        
        # Customer type validation
        customer_types = '"Regular,New,VIP,Student,Senior"'
        customer_validation = DataValidation(type="list", formula1=customer_types)
        customer_validation.add("H2:H101")
        ws.add_data_validation(customer_validation)
    
    def _add_borders(self, ws, cols: int, rows: int) -> None:
        """Add borders to the worksheet."""
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        for row in range(1, rows + 1):
            for col in range(1, cols + 1):
                ws.cell(row=row, column=col).border = thin_border
    
    def save_template(self) -> str:
        """Save the Excel template to file."""
        os.makedirs("excel_templates", exist_ok=True)
        self.workbook.save(self.template_path)
        return self.template_path


def main():
    """Create the Excel template."""
    print("Creating Daily Sales Sheet Excel Template...")
    
    creator = SalesSheetCreator()
    creator.create_sales_entry_sheet()
    creator.create_daily_summary_sheet()
    creator.create_product_analysis_sheet()
    creator.create_charts_sheet()
    
    template_path = creator.save_template()
    print(f"Excel template created successfully: {template_path}")
    print("\nTemplate includes:")
    print("- Sales Entry sheet with data validation")
    print("- Daily Summary with automatic calculations")
    print("- Product Analysis with performance metrics")
    print("- Charts & Reports sheet for visualizations")


if __name__ == "__main__":
    main()
